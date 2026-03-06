# -*- coding: utf-8 -*-
"""
IARM.py — импорт моделей и фонда защит из АРМ СРЗА (Excel .xlsx) в объекты mrtkz3.

Основные классы:
    ImpQ      -- узел сети (из таблицы «Наим.узлов»)
    ImpP      -- ветвь сети (из таблицы «Таблица ветвей»)
    ImpE      -- элемент сети (из таблицы «Наим.элементов»)
    ImpM      -- группа взаимоиндукций (из таблицы «Индуктивные группы»)
    ImpModel  -- контейнер для всех объектов; методы чтения и конвертации

Основные функции:
    parse_fond_xlsx(filename)           -- парсинг фонда защит (вертикальный формат)
    apply_fond_to_model(mdl, fond_list) -- привязка защит из фонда к объектам mrtkz3.Model
"""

import re
import numpy as np
import networkx as nx
import openpyxl


# ---------------------------------------------------------------------------
# Транслитерация
# ---------------------------------------------------------------------------

_TRANSLIT = {
    'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E', 'Ё': 'Ey',
    'Ж': 'Zh', 'З': 'Z', 'И': 'I', 'Й': 'Iy', 'К': 'K', 'Л': 'L', 'М': 'M',
    'Н': 'N', 'О': 'O', 'П': 'P', 'Р': 'R', 'С': 'S', 'Т': 'T', 'У': 'U',
    'Ф': 'F', 'Х': 'Kh', 'Ц': 'Tc', 'Ч': 'Ch', 'Ш': 'Sh', 'Щ': 'Shch',
    'Ы': 'Y', 'Э': 'Ee', 'Ю': 'Iu', 'Я': 'Ia',
}


def translite(s):
    res = ['q_']
    for ch in str(s):
        res.append(_TRANSLIT.get(ch, ch))
    return ''.join(res)


# ---------------------------------------------------------------------------
# Вспомогательные функции разбора значений ячеек
# ---------------------------------------------------------------------------

def _to_float(v):
    """Преобразует значение ячейки в float (поддерживает запятую как разделитель)."""
    if v is None:
        return 0.0
    if isinstance(v, str):
        v = v.strip().replace(',', '.')
        return float(v) if v else 0.0
    return float(v)


def _to_node_key(v):
    """Нормализует ключ узла: float → int, строка → stripped string."""
    if v is None:
        return 0
    if isinstance(v, float):
        return int(v)
    if isinstance(v, int):
        return v
    return v.strip()


# ---------------------------------------------------------------------------
# Классы промежуточного представления
# ---------------------------------------------------------------------------

class ImpQ:
    """Узел сети из таблицы «Наим.узлов» АРМ СРЗА."""

    def __init__(self, name, desc, key):
        self.name = name
        self.tlname = translite(name) if isinstance(name, str) else 'q_' + str(name)
        self.desc = desc.rstrip() if isinstance(desc, str) else (desc or '')
        self.key = key
        self.elem = None
        self.plist = []

    def addp(self, p):
        self.plist.append(p)


class ImpP:
    """Ветвь сети из таблицы «Таблица ветвей» АРМ СРЗА.

    Типы ветвей (self.typ):
        0   -- простая ветвь (линия)
        1   -- включённый выключатель
        101 -- отключённый выключатель (не импортируется в модель)
        3   -- трансформатор
        4   -- система / генератор (с ЭДС)
        5   -- ветвь с ёмкостной проводимостью B
    """

    def __init__(self, typ, par, q1, q2, Nel, Z1, Z2, Z0, EKB1, F1L, KB0):
        self.typ = typ
        self.par = par
        self.q1 = q1
        self.q2 = q2
        q1name = q1.name if isinstance(q1, ImpQ) else '0'
        q2name = q2.name if isinstance(q2, ImpQ) else '0'
        tlq1 = q1.tlname[2:] if isinstance(q1, ImpQ) else '0'
        tlq2 = q2.tlname[2:] if isinstance(q2, ImpQ) else '0'
        self.name = '{} {}-{}'.format(par, q1name, q2name)
        self.tlname = 'p_' + str(par) + '_' + tlq1 + '_' + tlq2
        self.Z1 = Z1
        self.Z2 = Z2
        self.Z0 = Z0
        self.EKB1 = EKB1
        self.F1L = F1L
        self.KB0 = KB0
        self.elem = Nel  # ImpE object


class ImpE:
    """Элемент сети из таблицы «Наим.элементов» АРМ СРЗА.

    Элемент — логическое объединение нескольких ветвей ImpP (линия, трансформатор и т.д.).
    """

    def __init__(self, name, desc):
        self.name = name
        self.desc = desc.rstrip() if isinstance(desc, str) else (desc or '')
        self.qlist = []
        self.plist = []

    def addp(self, p):
        self.plist.append(p)

    def addq(self, q):
        self.qlist.append(q)


class ImpM:
    """Группа взаимоиндукций нулевой последовательности."""

    def __init__(self, N):
        self.N = N
        self.M = np.zeros((N, N), dtype=np.cdouble)
        self.plist = []


# ---------------------------------------------------------------------------
# Основной класс модели
# ---------------------------------------------------------------------------

class ImpModel:
    """
    Контейнер для сетевой модели, импортированной из Excel АРМ СРЗА.

    Методы:
        ImpFromXLS(filename)  -- читает .xlsx файл с моделью
        to_mrtkz_model()      -- создаёт объект mrtkz3.Model напрямую
    """

    def __init__(self, name, desc):
        self.name = name
        self.desc = desc
        self.qlist = {}   # {node_key: ImpQ}
        self.plist = {}   # {'par q1-q2': ImpP}
        self.mlist = []   # [ImpM]
        self.elist = {}   # {elem_name: ImpE}
        self.el0 = ImpE(0, 'Общий элемент')
        self.elist[0] = self.el0

    # ------------------------------------------------------------------
    def ImpFromXLS(self, filename):
        """
        Читает файл модели АРМ СРЗА в формате .xlsx.

        Ожидаемые листы: «Наим.узлов», «Индуктивные группы»,
        «Таблица ветвей», «Наим.элементов».

        Параметры:
            filename -- путь к .xlsx файлу
        """
        wb = openpyxl.load_workbook(filename, data_only=True)

        QTable = wb['Наим.узлов']
        MTable = wb['Индуктивные группы']
        PTable = wb['Таблица ветвей']
        ETable = wb['Наим.элементов']

        # Счётчики из первой строки каждого листа
        NQ = int(re.search(r'\d+', str(QTable.cell(1, 1).value)).group())
        NP = int(re.search(r'\d+', str(PTable.cell(1, 1).value)).group())
        NM = int(re.search(r'\d+', str(MTable.cell(1, 1).value)).group())
        NE = int(re.search(r'\d+', str(ETable.cell(1, 1).value)).group())

        print(f'Кол-во узлов: {NQ}, ветвей: {NP}, инд.групп: {NM}, элементов: {NE}')

        # --- Элементы (строки 3..NE+2, т.к. строка 1 — заголовок, строка 2 — шапка) ---
        for row in range(3, NE + 3):
            el_name = _to_node_key(ETable.cell(row, 1).value)
            el_desc = ETable.cell(row, 2).value or ''
            ke = ImpE(el_name, el_desc)
            self.elist[el_name] = ke

        # --- Узлы ---
        for row in range(3, NQ + 3):
            q_name = _to_node_key(QTable.cell(row, 1).value)
            q_desc = QTable.cell(row, 2).value or ''
            q_key  = int(QTable.cell(row, 3).value or 0)
            kq = ImpQ(q_name, q_desc, q_key)
            self.qlist[q_name] = kq

        # --- Ветви ---
        for row in range(3, NP + 3):
            try:
                p_typ  = int(PTable.cell(row, 1).value or 0)
                p_par  = int(PTable.cell(row, 2).value or 0)
                p_q1k  = _to_node_key(PTable.cell(row, 3).value)
                p_q2k  = _to_node_key(PTable.cell(row, 4).value)
                p_elk  = _to_node_key(PTable.cell(row, 5).value)

                q1 = self.qlist.get(p_q1k, 0) if p_q1k != 0 else 0
                q2 = self.qlist.get(p_q2k, 0) if p_q2k != 0 else 0
                el = self.elist.get(p_elk, self.el0)

                Z1  = _to_float(PTable.cell(row, 6).value) + 1j * _to_float(PTable.cell(row, 7).value)
                EKB1 = _to_float(PTable.cell(row, 8).value)
                F1L  = _to_float(PTable.cell(row, 9).value)
                Z0   = _to_float(PTable.cell(row, 10).value) + 1j * _to_float(PTable.cell(row, 11).value)
                KB0  = _to_float(PTable.cell(row, 12).value)
                Z2   = _to_float(PTable.cell(row, 13).value) + 1j * _to_float(PTable.cell(row, 14).value)

                if Z2 == 0:
                    Z2 = Z1
                if p_typ not in (1, 101) and Z0 == 0:
                    Z0 = Z1

                kp = ImpP(p_typ, p_par, q1, q2, el, Z1, Z2, Z0, EKB1, F1L, KB0)
                el.addp(kp)
                if isinstance(q1, ImpQ):
                    q1.addp(kp)
                if isinstance(q2, ImpQ):
                    q2.addp(kp)
                self.plist[kp.name] = kp
            except Exception as e:
                print(f'Ошибка при чтении ветви в строке {row}: {e}')

        # Привязка узлов к элементам
        for kq in self.qlist.values():
            if not kq.plist:
                continue
            kq.elem = kq.plist[0].elem
            for kp in kq.plist[1:]:
                if kq.elem is not kp.elem:
                    kq.elem = self.el0
                    self.el0.addq(kq)
                    break
            else:
                kq.elem.addq(kq)

        # --- Взаимоиндукции ---
        rows_iter = iter(range(2, MTable.max_row + 1))
        for _ in range(NM):
            try:
                ijk = next(rows_iter)
                header_val = str(MTable.cell(ijk, 1).value or '')
                km_n = int(re.search(r'\d+', header_val).group())
                km = ImpM(km_n)
                next(rows_iter)  # строка с заголовком столбцов
                for i in range(km_n):
                    data_row = next(rows_iter)
                    p_par = int(MTable.cell(data_row, 1).value or 0)
                    p_q1k = _to_node_key(MTable.cell(data_row, 2).value)
                    p_q2k = _to_node_key(MTable.cell(data_row, 3).value)
                    key = '{} {}-{}'.format(p_par, p_q1k, p_q2k)
                    if key in self.plist:
                        km.plist.append(self.plist[key])
                    else:
                        km.plist.append(None)
                    for j in range(km_n):
                        col = 4 + 2 * j
                        km.M[i, j] = (_to_float(MTable.cell(data_row, col).value) +
                                      1j * _to_float(MTable.cell(data_row, col + 1).value))
                self.mlist.append(km)
            except StopIteration:
                break
            except Exception as e:
                print(f'Ошибка при чтении индуктивной группы: {e}')

    # ------------------------------------------------------------------
    def to_mrtkz_model(self):
        """
        Создаёт объект mrtkz3.Model напрямую из данных ImpModel.

        Возвращает:
            mrtkz3.Model -- готовая расчётная модель с узлами, ветвями,
                            элементами и взаимоиндукциями
        """
        import mrtkz3 as mrtkz

        mdl = mrtkz.Model(desc=self.desc)
        q_map = {}   # ImpQ.name → mrtkz.Q
        e_map = {}   # ImpE.name → mrtkz.Element
        p_map = {}   # ImpP.name → mrtkz.P

        # Узлы
        for imp_q in self.qlist.values():
            q = mrtkz.Q(mdl, str(imp_q.name), desc=str(imp_q.desc or ''))
            q_map[imp_q.name] = q

        # Элементы (кроме служебного el0)
        for ename, imp_e in self.elist.items():
            if ename == 0:
                continue
            e = mrtkz.Element(mdl, imp_e.name, desc=str(imp_e.desc or ''))
            e_map[ename] = e

        # Ветви
        for imp_p in self.plist.values():
            if imp_p.typ == 101:   # отключённый выключатель — пропустить
                continue

            q1 = q_map.get(imp_p.q1.name) if isinstance(imp_p.q1, ImpQ) else 0
            q2 = q_map.get(imp_p.q2.name) if isinstance(imp_p.q2, ImpQ) else 0
            Z  = (imp_p.Z1, imp_p.Z2, imp_p.Z0)

            try:
                if imp_p.typ in (0, 1):
                    p = mrtkz.P(mdl, imp_p.name, q1, q2, Z)
                elif imp_p.typ == 3:    # трансформатор
                    p = mrtkz.P(mdl, imp_p.name, q1, q2, Z, T=(imp_p.EKB1, 0))
                elif imp_p.typ == 4:    # система / генератор
                    E_mag = imp_p.EKB1 * 1000 / 1.732
                    E_val = E_mag * np.exp(1j * np.pi / 180 * imp_p.F1L)
                    p = mrtkz.P(mdl, imp_p.name, q1, q2, Z, E=(E_val, 0, 0))
                elif imp_p.typ == 5:    # ветвь с ёмкостной проводимостью
                    B = (imp_p.EKB1 * 1e-6j, imp_p.EKB1 * 1e-6j, imp_p.KB0 * 1e-6j)
                    p = mrtkz.P(mdl, imp_p.name, q1, q2, Z, B=B)
                else:
                    continue
            except Exception as e:
                print(f'Ошибка при создании ветви {imp_p.name}: {e}')
                continue

            # Привязываем к элементу
            if isinstance(imp_p.elem, ImpE) and imp_p.elem.name in e_map:
                e_map[imp_p.elem.name].addp(p)

            p_map[imp_p.name] = p

        # Взаимоиндукции
        for imp_m in self.mlist:
            for i in range(imp_m.N):
                for j in range(i):
                    pi = p_map.get(imp_m.plist[i].name) if imp_m.plist[i] else None
                    pj = p_map.get(imp_m.plist[j].name) if imp_m.plist[j] else None
                    if pi and pj:
                        mrtkz.M(mdl,
                                f'{pi.name} -- {pj.name}',
                                pi, pj,
                                imp_m.M[i, j], imp_m.M[j, i])

        return mdl


# ---------------------------------------------------------------------------
# Парсер фонда защит
# ---------------------------------------------------------------------------

def parse_fond_xlsx(filename):
    """
    Парсит файл фонда защит АРМ СРЗА в вертикальном формате (.xlsx).

    Для каждой защиты (комплект 1) с параметрами срабатывания возвращает dict:
        prot_num   -- номер защиты (int, напр. 21)
        elem_num   -- номер элемента (= prot_num // 10)
        node_num   -- номер узла внутри элемента (= prot_num % 10; 1=q1, 2=q2)
        vl         -- название ВЛ/КЛ
        ps         -- название ПС / узла
        ktt        -- коэффициент ТТ (строка, напр. '1000/5')
        kth        -- коэффициент ТН (строка, напр. '1100')
        panel_type -- тип панели (напр. 'ТЗНП')
        relay_type -- тип реле мощности (напр. 'ШЭ', 'ЭЛ/МЕХ')
        stages     -- список dict [{'stage': int, 'I0': float, 't': float,
                                    'ang': float, 'direction': str}]
        warnings   -- список строк с замечаниями (напр. о нескольких активных комплектах)

    Правила:
        - Импортируется только комплект 1.
        - Если несколько комплектов имеют параметры → берётся первый,
          в поле warnings записывается замечание.
        - Если комплект 1 не имеет параметров → защита пропускается
          (даже если комплект 2 имеет параметры).
        - Ступени с I0=0 и t=0 считаются несуществующими и не импортируются.
        - Импортируются только защиты с типом панели 'ТЗНП'.

    Параметры:
        filename -- путь к .xlsx файлу фонда защит

    Возвращает:
        list[dict]
    """
    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.active

    # Читаем все строки в память
    all_rows = []
    for r in range(1, ws.max_row + 1):
        row = tuple(ws.cell(r, c).value for c in range(1, ws.max_column + 1))
        all_rows.append(row)

    results = []
    i = 0
    n = len(all_rows)

    while i < n:
        cell0 = str(all_rows[i][0] or '').strip()

        if not cell0.startswith('Защита'):
            i += 1
            continue

        # ---- Новый заголовок защиты ----
        m = re.search(r'\d+', cell0)
        if not m:
            i += 1
            continue
        prot_num = int(m.group())
        elem_num = prot_num // 10
        node_num = prot_num % 10

        block = {
            'prot_num':   prot_num,
            'elem_num':   elem_num,
            'node_num':   node_num,
            'vl':         '',
            'ps':         '',
            'ktt':        '',
            'kth':        '',
            'panel_type': '',
            'relay_type': '',
            'stages':     [],
            'warnings':   [],
        }
        i += 1

        # --- Сбор всех комплектов для этой защиты ---
        # Каждый комплект: {'komplekt': int, 'has_params': bool, 'stages': list}
        # Заголовочные поля (vl, ps, ktt, kth, ...) берём из первого вхождения
        komplekty = []
        cur_k = {'komplekt': 1, 'has_params': True, 'stages': [], 'header_read': False}

        while i < n:
            r0 = str(all_rows[i][0] or '').strip()

            # Новый блок защиты — прерываем
            if r0.startswith('Защита'):
                break

            # --- Заголовочные поля ---
            if r0.startswith('ВЛ'):
                if not cur_k['header_read']:
                    block['vl'] = r0[3:].strip()
            elif r0.startswith('ПС') or r0.startswith('Узел'):
                if not cur_k['header_read']:
                    block['ps'] = r0.split(None, 1)[-1].strip()
            elif r0.startswith('KTT'):
                if not cur_k['header_read']:
                    block['ktt'] = r0[4:].strip()
            elif r0.startswith('KTH'):
                if not cur_k['header_read']:
                    block['kth'] = r0[4:].strip()
            elif r0.startswith('Тип панели'):
                if not cur_k['header_read']:
                    block['panel_type'] = r0.split()[-1]
            elif r0.startswith('Тип реле'):
                if not cur_k['header_read']:
                    block['relay_type'] = r0.split()[-1]
                cur_k['header_read'] = True

            # Начало нового комплекта
            elif r0.startswith('комплект') or r0.startswith('Комплект'):
                km = re.search(r'\d+', r0)
                k_num = int(km.group()) if km else 1
                komplekty.append(cur_k)
                cur_k = {'komplekt': k_num, 'has_params': True, 'stages': [], 'header_read': True}

            # Нет параметров срабатывания
            elif r0.startswith('НЕТ ПАРАМЕТРОВ'):
                cur_k['has_params'] = False
                i += 1
                break

            # Таблица параметров
            elif r0.startswith('Параметры срабатывания'):
                vals_header = all_rows[i][1:]
                n_stages = sum(1 for v in vals_header if v is not None and 'ступ' in str(v))
                stages = [{'stage': s + 1, 'I0': 0.0, 't': 0.0, 'ang': 0.0, 'direction': 'ненапр'}
                          for s in range(n_stages)]
                i += 1

                while i < n:
                    r0 = str(all_rows[i][0] or '').strip()
                    if (r0.startswith('Защита') or
                            r0.startswith('комплект') or r0.startswith('Комплект') or
                            r0.startswith('НЕТ ПАРАМЕТРОВ')):
                        break
                    vals = all_rows[i][1:]
                    if r0 == 'Ток срабатывания':
                        for s in range(n_stages):
                            stages[s]['I0'] = _to_float(vals[s] if s < len(vals) else None)
                    elif r0 == 'Угол макс чувс ОНМ':
                        for s in range(n_stages):
                            stages[s]['ang'] = _to_float(vals[s] if s < len(vals) else None)
                    elif r0 == 'Время срабатывания':
                        for s in range(n_stages):
                            stages[s]['t'] = _to_float(vals[s] if s < len(vals) else None)
                    elif r0 == 'Направление действия':
                        for s in range(n_stages):
                            v = vals[s] if s < len(vals) else None
                            stages[s]['direction'] = str(v or 'ненапр').strip()
                        i += 1
                        break
                    i += 1
                cur_k['stages'] = stages
                continue  # не делаем i += 1 в конце — уже сдвинули

            i += 1

        komplekty.append(cur_k)

        # --- Выбор комплекта ---
        # Берём только те, у которых has_params=True
        active = [k for k in komplekty if k['has_params']]

        if not active:
            continue  # нет ни одного активного комплекта

        # Берём комплект 1, или первый активный, если комплект 1 не имеет параметров
        k1 = next((k for k in active if k['komplekt'] == 1), None)
        if k1 is None:
            continue  # комплект 1 без параметров → пропустить всю защиту

        if len(active) > 1:
            other_nums = [k['komplekt'] for k in active if k['komplekt'] != 1]
            block['warnings'].append(
                f'Защита {prot_num}: несколько активных комплектов {[k["komplekt"] for k in active]}; '
                f'импортирован только комплект 1, комплекты {other_nums} проигнорированы.'
            )

        # Фильтруем ступени: оставляем только с I0>0 или t>0
        block['stages'] = [s for s in k1['stages'] if s['I0'] > 0 or s['t'] > 0]

        if block['stages'] and block['panel_type'] == 'ТЗНП':
            results.append(block)

    return results


# ---------------------------------------------------------------------------
# Применение фонда к модели
# ---------------------------------------------------------------------------

def apply_fond_to_model(mdl, fond_list):
    """
    Создаёт объекты mrtkz3.protection на ветвях модели по данным из фонда.

    Маппинг:
        fond['elem_num'] → Element в mdl.be (по атрибуту name)
        fond['node_num'] == 1  → q1 первой ветви элемента
        fond['node_num'] == 2  → q2 последней ветви элемента

    Параметры:
        mdl       -- mrtkz3.Model с загруженными элементами
        fond_list -- результат parse_fond_xlsx()

    Возвращает:
        int -- количество успешно созданных объектов protection
    """
    import mrtkz3 as mrtkz

    # Словарь элементов по номеру (name хранится как str)
    elem_by_num = {e.name: e for e in mdl.be}

    created = 0
    for fond in fond_list:
        elem = elem_by_num.get(str(fond['elem_num']))
        if elem is None or not elem.plist:
            continue

        # Определяем ветвь и узел
        if fond['node_num'] == 1:
            branch = elem.plist[0]
            q = branch.q1
        else:
            branch = elem.plist[-1]
            q = branch.q2

        if q == 0:
            continue

        for s in fond['stages']:
            direction = s.get('direction', 'ненапр')
            rnm_on = (direction != 'ненапр')
            ang = s.get('ang', 0.0)
            I0 = s['I0'] if s['I0'] > 0 else 999999

            mrtkz.protection(
                p=branch,
                q=q,
                stage=s['stage'],
                I0=I0,
                t=s['t'],
                type='ТЗНП',
                P_rnm=0,
                rnm_base_angle=ang,
                rnm_on=rnm_on,
                stage_on=True,
                ktt=fond.get('ktt', ''),
                kth=fond.get('kth', ''),
                relay_type=fond.get('relay_type', ''),
            )
            created += 1

    return created
